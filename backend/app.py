from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
import os
from datetime import datetime, timedelta
from functools import wraps

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-in-production')
EXCEL_FILE = 'users_db.xlsx'
INITIAL_WALLET = 100000

# ===== EXCEL DATABASE FUNCTIONS =====

def init_excel_db():
    """Initialize Excel database if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        
        # Users sheet
        ws_users = wb.active
        ws_users.title = 'Users'
        ws_users.append(['Username', 'Email', 'Password', 'Wallet Balance', 'Total Invested', 'Created At'])
        
        # Transactions sheet
        ws_transactions = wb.create_sheet('Transactions')
        ws_transactions.append(['Username', 'Action', 'Stock Symbol', 'Quantity', 'Price', 'Total', 'Timestamp'])
        
        # Portfolio sheet
        ws_portfolio = wb.create_sheet('Portfolio')
        ws_portfolio.append(['Username', 'Stock Symbol', 'Quantity', 'Avg Buy Price', 'Current Value'])
        
        wb.save(EXCEL_FILE)

def get_user_row(username):
    """Find user row in Users sheet"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Users']
    
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == username:
            return row
    return None

def user_exists(username):
    """Check if user exists"""
    return get_user_row(username) is not None

def get_user_data(username):
    """Get user data"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Users']
    row = get_user_row(username)
    
    if row:
        return {
            'username': ws[f'A{row}'].value,
            'email': ws[f'B{row}'].value,
            'wallet_balance': ws[f'D{row}'].value,
            'total_invested': ws[f'E{row}'].value,
        }
    return None

def add_transaction(username, action, symbol, quantity, price):
    """Add transaction to Transactions sheet"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Transactions']
    
    total = quantity * price
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    ws.append([username, action, symbol, quantity, price, total, timestamp])
    wb.save(EXCEL_FILE)

def update_portfolio(username, symbol, quantity, action, price):
    """Update or create portfolio entry"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Portfolio']
    
    # Find existing portfolio entry
    portfolio_row = None
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == username and ws[f'B{row}'].value == symbol:
            portfolio_row = row
            break
    
    if action == 'BUY':
        if portfolio_row:
            # Update existing position
            current_qty = ws[f'C{portfolio_row}'].value
            current_avg_price = ws[f'D{portfolio_row}'].value
            new_qty = current_qty + quantity
            new_avg_price = ((current_qty * current_avg_price) + (quantity * price)) / new_qty
            
            ws[f'C{portfolio_row}'].value = new_qty
            ws[f'D{portfolio_row}'].value = new_avg_price
            ws[f'E{portfolio_row}'].value = new_qty * price
        else:
            # Create new position
            ws.append([username, symbol, quantity, price, quantity * price])
    
    elif action == 'SELL':
        if portfolio_row:
            current_qty = ws[f'C{portfolio_row}'].value
            new_qty = current_qty - quantity
            
            if new_qty <= 0:
                # Delete row if quantity becomes 0
                ws.delete_rows(portfolio_row, 1)
            else:
                ws[f'C{portfolio_row}'].value = new_qty
                ws[f'E{portfolio_row}'].value = new_qty * price
    
    wb.save(EXCEL_FILE)

def update_wallet(username, amount):
    """Update wallet balance"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Users']
    row = get_user_row(username)
    
    if row:
        current_balance = ws[f'D{row}'].value or INITIAL_WALLET
        total_invested = ws[f'E{row}'].value or 0
        
        ws[f'D{row}'].value = current_balance + amount
        ws[f'E{row}'].value = total_invested + abs(amount)
        
        wb.save(EXCEL_FILE)
        return True
    return False

# ===== JWT HELPER FUNCTIONS =====

def create_token(username):
    """Create JWT token"""
    payload = {
        'username': username,
        'exp': datetime.utcnow() + timedelta(hours=24)
    }
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')

def verify_token(f):
    """Verify JWT token decorator"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'error': 'Missing token'}), 401
        
        try:
            token = token.split(' ')[1]  # Remove 'Bearer '
            payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            username = payload['username']
        except Exception as e:
            return jsonify({'error': 'Invalid token'}), 401
        
        return f(username, *args, **kwargs)
    
    return decorated_function

# ===== AUTH ENDPOINTS =====

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register new user"""
    data = request.json
    username = data.get('username', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    
    if not username or not email or not password:
        return jsonify({'error': 'Missing fields'}), 400
    
    if user_exists(username):
        return jsonify({'error': 'Username already exists'}), 409
    
    # Add user to Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Users']
    
    hashed_password = generate_password_hash(password)
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    ws.append([username, email, hashed_password, INITIAL_WALLET, 0, created_at])
    wb.save(EXCEL_FILE)
    
    token = create_token(username)
    
    return jsonify({
        'message': 'Registration successful',
        'token': token,
        'username': username,
        'wallet': INITIAL_WALLET
    }), 201

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login user"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'error': 'Missing credentials'}), 400
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Users']
    row = get_user_row(username)
    
    if not row:
        return jsonify({'error': 'User not found'}), 404
    
    stored_password = ws[f'C{row}'].value
    
    if not check_password_hash(stored_password, password):
        return jsonify({'error': 'Invalid password'}), 401
    
    token = create_token(username)
    user_data = get_user_data(username)
    
    return jsonify({
        'message': 'Login successful',
        'token': token,
        'username': username,
        'email': user_data['email'],
        'wallet': user_data['wallet_balance'],
        'total_invested': user_data['total_invested']
    }), 200

# ===== USER ENDPOINTS =====

@app.route('/api/user/profile', methods=['GET'])
@verify_token
def get_profile(username):
    """Get user profile"""
    user_data = get_user_data(username)
    
    if not user_data:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify(user_data), 200

@app.route('/api/user/portfolio', methods=['GET'])
@verify_token
def get_portfolio(username):
    """Get user portfolio"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Portfolio']
    
    portfolio = []
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == username:
            portfolio.append({
                'symbol': ws[f'B{row}'].value,
                'quantity': ws[f'C{row}'].value,
                'avg_buy_price': ws[f'D{row}'].value,
                'current_value': ws[f'E{row}'].value,
            })
    
    return jsonify({'portfolio': portfolio}), 200

# ===== TRADING ENDPOINTS =====

@app.route('/api/trade/buy', methods=['POST'])
@verify_token
def buy_stock(username):
    """Buy stock"""
    data = request.json
    symbol = data.get('symbol', '').upper()
    quantity = int(data.get('quantity', 0))
    price = float(data.get('price', 0))
    
    if not symbol or quantity <= 0 or price <= 0:
        return jsonify({'error': 'Invalid trade parameters'}), 400
    
    user_data = get_user_data(username)
    total_cost = quantity * price
    
    if user_data['wallet_balance'] < total_cost:
        return jsonify({'error': 'Insufficient funds'}), 400
    
    # Execute trade
    update_wallet(username, -total_cost)
    update_portfolio(username, symbol, quantity, 'BUY', price)
    add_transaction(username, 'BUY', symbol, quantity, price)
    
    return jsonify({
        'message': 'Purchase successful',
        'symbol': symbol,
        'quantity': quantity,
        'total_cost': total_cost
    }), 200

@app.route('/api/trade/sell', methods=['POST'])
@verify_token
def sell_stock(username):
    """Sell stock"""
    data = request.json
    symbol = data.get('symbol', '').upper()
    quantity = int(data.get('quantity', 0))
    price = float(data.get('price', 0))
    
    if not symbol or quantity <= 0 or price <= 0:
        return jsonify({'error': 'Invalid trade parameters'}), 400
    
    # Check if user has enough stocks
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Portfolio']
    
    portfolio_row = None
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == username and ws[f'B{row}'].value == symbol:
            portfolio_row = row
            break
    
    if not portfolio_row or ws[f'C{portfolio_row}'].value < quantity:
        return jsonify({'error': 'Insufficient stocks to sell'}), 400
    
    # Execute trade
    total_proceeds = quantity * price
    update_wallet(username, total_proceeds)
    update_portfolio(username, symbol, quantity, 'SELL', price)
    add_transaction(username, 'SELL', symbol, quantity, price)
    
    return jsonify({
        'message': 'Sale successful',
        'symbol': symbol,
        'quantity': quantity,
        'total_proceeds': total_proceeds
    }), 200

@app.route('/api/user/transactions', methods=['GET'])
@verify_token
def get_transactions(username):
    """Get user transactions"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Transactions']
    
    transactions = []
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == username:
            transactions.append({
                'action': ws[f'B{row}'].value,
                'symbol': ws[f'C{row}'].value,
                'quantity': ws[f'D{row}'].value,
                'price': ws[f'E{row}'].value,
                'total': ws[f'F{row}'].value,
                'timestamp': ws[f'G{row}'].value,
            })
    
    return jsonify({'transactions': list(reversed(transactions))}), 200

# ===== UTILITY ENDPOINTS =====

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'ok'}), 200

if __name__ == '__main__':
    init_excel_db()
    app.run(debug=True, port=5000)
